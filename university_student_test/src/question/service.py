from sqlalchemy.ext.asyncio import AsyncSession
from core.utils.basic_service import BasicService
from .schemas import QuestionCreate, QuestionUpdate, QuestionBase
from sqlalchemy import insert, select
from openpyxl import load_workbook
from core.models.questions import Question
from fastapi import HTTPException, status, UploadFile
import tempfile


class QuestionService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.service = BasicService(db=session)

    async def create_question(self, question_data: QuestionCreate):
        return await self.service.create(model=Question, obj_items=question_data)

    async def create_question_by_exel(self, subject_id: int, file: UploadFile, user_id: int):
        # Validate file extension
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid file format. Please upload an Excel file."
            )

        # Save uploaded file to a temp file
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            temp_path = tmp.name

        # Load Excel workbook
        wb = load_workbook(temp_path)
        sheet = wb.active

        values = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # first column = question text
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Row {idx} is invalid: 'text' column is empty"
                )

            values.append({
                "subject_id": subject_id,
                "user_id": user_id,
                "text": row[0],
                "option_a": row[1],
                "option_b": row[2],
                "option_c": row[3],
                "option_d": row[4],
            })

        # Bulk insert
        stmt = insert(Question).values(values)
        await self.session.execute(stmt)
        await self.session.commit()

        return {"status": "success", "message": f"{len(values)} questions uploaded successfully"}
    
    
    async def get_question_by_id(
        self, 
        question_id: int, 
        user_id: int, 
        role: str | None = None
        ):
        
        if role == "admin":
            stmt = select(Question).where(Question.id == question_id)
        else:
            stmt = select(Question).where(
                Question.id == question_id, 
                Question.user_id == user_id
                )
            
        result = await self.session.execute(stmt)
        question_data = result.scalars().first()
        
        if not question_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Question not found"
            )
            
        return question_data

    async def get_all_question(
        self,
        is_admin: str, 
        user_id: int,
        limit: int = 20, 
        offset: int = 0
    ):
        stmt = select(Question)

        # Apply condition only if user is not admin
        if is_admin != "admin":
            stmt = stmt.where(Question.user_id == user_id)
            

        # Apply pagination correctly
        stmt = stmt.limit(limit).offset(offset)

        result = await self.session.execute(stmt)
        questions_data = result.scalars().all()
        
        if not questions_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Questions not found"
            )
        return questions_data


    async def update_question(
        self,
        question_id: int,
        question_data: "QuestionUpdate",
        user_id: int,
        is_admin: str | None,
    ):
        # Get the question
        stmt = select(Question).where(Question.id == question_id)
        result = await self.session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Question not found"
            )

        # Check permission — only admins or the question's owner can update
        if is_admin != "admin" and question.user_id != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not allowed to update this question"
            )

        # Apply updates from QuestionUpdate model
        for field, value in question_data.model_dump(exclude_unset=True).items():
            setattr(question, field, value)

        await self.session.commit()
        await self.session.refresh(question)
        return question


    async def delete_question(
            self,
            question_id: int,
            user_id: int,
            is_admin: str | None
        ):
            # Check if question exists
            stmt = select(Question).where(Question.id == question_id)
            result = await self.session.execute(stmt)
            question = result.scalar_one_or_none()

            if not question:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Question not found"
                )

            # Check permission: allow only admin or owner
            if is_admin != "admin" and question.user_id != user_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You are not allowed to delete this question"
                )

            # Delete the question
            await self.session.delete(question)
            await self.session.commit()

            return {"message": "Deleted successfully"}
        
    # async def check_by_teacher_id(
    #     self,  
    #     limit: int | None = None,
    #     offset: int | None = 0,
    #     is_all: bool = False,
    #     filters: list | None = None
    # ):
    #     # Ensure filters is a list
    #     if filters is None:
    #         filters = []

    #     # Start base query (you can adjust the model)
    #     stmt = select(Question)

    #     # Apply filters if provided
    #     for condition in filters:
    #         stmt = stmt.where(condition)

    #     # Apply pagination
    #     if limit is not None:
    #         stmt = stmt.limit(limit)
    #     if offset:
    #         stmt = stmt.offset(offset)

    #     # Execute
    #     result = await self.session.execute(stmt)
    #     data = result.scalars().all()

    #     # Handle empty result
    #     if not data:
    #         raise HTTPException(
    #             status_code=status.HTTP_400_BAD_REQUEST,
    #             detail="Question not found"
    #         )

    #     return data
    
    

